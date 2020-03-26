# tydzien = ['pon', 'wt', 'śr', 'czw', 'pia', 'sob', 'nie']
# weekend = ['sob', 'nie']
#
#
# def wyswietl(text):
#     print(text)
#
# for dzien in tydzien:
#     if dzien in weekend:
#         print('Jest weekend!')
#         wyswietl('Jest weekend!')

# def fun1():
#     print('fun1')
#
# def fun2():
#     print('fun2')
#
# def fun3():
#     print('fun3')
#
# def wyjscie():
#     print('Nie to, bajo!')
#     exit()
#
# print('Hej, Multitool!'
#       '\n1 - func 1'
#       '\n2 - func 2'
#       '\n3 - func 3'
#       '\nR - random'
#       '\nX - exit')
#
# program = input('Co chcesz zrobic?')
# funkcje = {'1':{'nazwa':, 'Funkcja 1', 'call': fun1},
#             '2':{fun2, '3':fun3, 'X': exit}
# try:
#     funkcje[program]()
# except:
#     print('Nie ma takiego programu')
import  openpyxl


# excel.save('test.xlsx')
# excel = openpyxl.load_workbook('test.xlsx')
# arkusz = excel.active
# komorka = arkusz.cell(1,1)
# komorka.value='1'
# print(komorka.value)
# # excel.save('test.xlsx')

def tworz_tabliczka(ilosc=10):
    zakres = [1,ilosc]
    excel = openpyxl.load_workbook('test.xlsx')
    nazwa_pliku = 'test.xlsx'
    arkusz = excel.active
    for row in range(1, ilosc +1):
        for col in range(1, ilosc + 1):
            komorka = arkusz.cell(row, col)
            komorka.value = row * col
    excel.save('test.xlsx')

tworz_tabliczka(10)






# excel = openpyxl.load_workbook('test.xlsx')
# arkusz = excel.active
# arkusze = excel.sheetnames
# arkusz = excel['Arkusz1']
# komorka = arkusz.cell(row = 1, column = 1)
# print(komorka.value)
# komorka.value = 'Nowa wartość'
# excel.save('plik.xlsx')

